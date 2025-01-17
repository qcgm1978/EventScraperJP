import requests
import re
from bs4 import BeautifulSoup
import openpyxl
import pykakasi
import pandas as pd
import time
import os
from openpyxl.styles import Font, Color, PatternFill, Alignment, Fill
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
EXCEL_FILE = "EventsJP2025.xlsx"

def save_workbook(workbook):   
    try:
        workbook.save(EXCEL_FILE)
    except: 
        print("Close Excel you imbecile!")
        time.sleep(15)
        save_workbook(workbook)
        
def convert_to_romaji(japanese_text):
    kks = pykakasi.kakasi()
    result = kks.convert(japanese_text)
    romaji_text = ''.join([item['hepburn'] for item in result])
    return romaji_text

def doc_from_url(url):
    pagePia = requests.get(url)
    pagePia.encoding = 'utf-8'
    html_content_Pia = pagePia.text
    return BeautifulSoup(html_content_Pia, "html.parser")

def OpenSheet(sheet_name, header):
    if os.path.exists(EXCEL_FILE):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    else:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(header) 
    sheet = workbook[sheet_name]
    return workbook, sheet

def style_sort_excel(sheet_name, sorting_column):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    column_count = sheet.max_column
    
    df = pd.DataFrame(sheet.values)
    headers = df.iloc[0]
    df = df[1:]
    df.columns = headers
    df = df.sort_values(by=sorting_column, ascending=True)
    sheet.delete_rows(1, sheet.max_row)
    for row in [df.columns.tolist()] + df.values.tolist():
       sheet.append(row)
    
    title_row_style = Font(size=14, color="FFFFFF", bold=True)
    for i in range (0,column_count):
        sheet.cell(row=1, column=i+1).font = title_row_style
    dim_holder = DimensionHolder(worksheet=sheet)
    for col in range(sheet.min_column, sheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=35)
    sheet.column_dimensions = dim_holder
    
    for z in range (0, column_count):
        sheet.cell(row=1, column = z + 1).fill = PatternFill(start_color="38AA49", end_color="38AA49", fill_type="solid")
    for x in range(2, row_count):
        for z in range (0, column_count): 
            c = sheet.cell(row=x, column=z + 1)
            if x % 2 != 0:
                c.fill = PatternFill(start_color="ACFFB8", end_color="ACFFB8", fill_type="solid")
    if row_count % 2 != 0:
        for z in range (0, column_count):
            l = sheet.cell(row=row_count, column=z + 1)
            l.fill = PatternFill(start_color="ACFFB8", end_color="ACFFB8", fill_type="solid")