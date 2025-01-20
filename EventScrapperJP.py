import requests
import re
from bs4 import BeautifulSoup
import openpyxl
import pykakasi
import pandas as pd
import time
import os
import concurrent.futures
import random
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from openpyxl.styles import Font, Color, PatternFill, Alignment, Fill
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

#If the front end allows it, the user might choose their own BASE_FOLDER. Then you should check if the BASE_FOLDER was chosen, if not, use the default one. 
#Make the correct function and set up the deault BASE_FOLDER as rf"C:\Users\{username}\Documents\EventScrapperJP" 
username=os.getlogin()
BASE_FOLDER = rf"C:\Users\{username}\Documents\EventScrapperJP" 
EXCEL_FILE = rf"{BASE_FOLDER}\EventsJP2025.xlsx"
HEADER = ["Name", "Romaji", "Place", "Beginning Date", "Ending Date", "Link"]

def save_workbook(workbook):    
    try:
        workbook.save(EXCEL_FILE)
    except: 
        print("Close Excel you imbecile!")
        time.sleep(15)
        save_workbook(workbook)

def convert_to_romaji(japanese_text):
    kks = pykakasi.kakasi()
    result = kks.convert(japanese_text)
    romaji_text = ''.join([item['hepburn'] for item in result])
    return romaji_text

def doc_from_url(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0",
        "Accept-Encoding": "gzip, deflate, br, zstd",
        "Accept-Language": "pl,en;q=0.9,en-GB;q=0.8,en-US;q=0.7",
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7"
    }
    print(f"Scraping page: {url}")
    while True:
        try:
            page = requests.get(url, headers=headers)
            page.encoding = 'utf-8'
            html_content_Pia = page.text
            print("Page scraped successfully.")
            return BeautifulSoup(html_content_Pia, "html.parser")
        except:
            print("Error with the page. Retrying...")
            time.sleep(random.randint(20, 60))

def PiaInnerScrapper(url):
    Inner_doc_Pia = doc_from_url(url)
    for div_Inner_Pia in Inner_doc_Pia.find_all("div", class_="textDefinitionList-2024__item"):
        dt_tag_Pia = div_Inner_Pia.find("dt", class_="textDefinitionList-2024__title")
        if dt_tag_Pia and "会場" in dt_tag_Pia.get_text(strip=True):
            dd_tag_Pia = div_Inner_Pia.find("dd", class_="textDefinitionList-2024__desc")
            if dd_tag_Pia:
                return dd_tag_Pia.get_text(strip=True)
                #if PPia:
                #    return convert_to_romaji(PPia)
                #else:
                #    return None   

def PiaScrapper(doc_Pia):
    i=0
    Piaconcerts = []
    for div_Pia in doc_Pia.find_all("div"):
        a_tag_Pia = div_Pia.find("a")
        
        if a_tag_Pia:
            figcaption_Pia = a_tag_Pia.find("figcaption")
            
            if figcaption_Pia:
                
                name_Pia = figcaption_Pia.find("h2")
                namePia = name_Pia.get_text(strip=True) if name_Pia else None
                
                if namePia:
                    romajiPia = convert_to_romaji(namePia)
                else:
                    romajiPia = None
                    
                date_Pia= figcaption_Pia.find("span")
                datePia = date_Pia.get_text(strip=True) if date_Pia else None
                
                linkPia = a_tag_Pia.get("href") if a_tag_Pia else None
                               
                placePia = PiaInnerScrapper(linkPia)
                                 
                if namePia and datePia and linkPia:
                    Piaconcerts.append({"Name": namePia, "Romaji": romajiPia, "Place": placePia, "Date": datePia, "Link": linkPia})
                    i+=1
                    #if i>1:
                    #    break #tester
                    #print(i)                       
    print(f"Finished scraping current site. Proceeding to the next one.")
    return Piaconcerts

def eplusScrapper(month):
    i=0
    Eplusconcerts = []
    url = f"https://eplus.jp/sf/event/month-0{month}"
    doc_eplus = doc_from_url(url)
    
    li_eplus = doc_eplus.find("li", class_="block-paginator__item block-paginator__item--last")
    if li_eplus:
        max_pages = int(li_eplus.get_text(strip=True))
        print(f"Max pages: {max_pages}")
    
        for page in range(1, max_pages+1):
            url = f"https://eplus.jp/sf/event/month-0{month}/p{page}"
            print(f"Scraping page: {page}")
            
            retries = 5
            while retries > 0:
                try:
                    doc_eplus = doc_from_url(url)
                    ticket_div = doc_eplus.find("div", class_="block-ticket-list__content output")
                    tickets = ticket_div.find_all("a")

                    for ticket in tickets:
                        nameEplus = (ticket.find("h3", class_="ticket-item__title")).get_text(strip=True)
                        
                        if nameEplus:
                            romajiEplus = convert_to_romaji(nameEplus)
                        else:
                            romajiEplus = None
                        
                        date_year = ticket.find_all("span", class_="ticket-item__yyyy") 
                        date_mmdd = ticket.find_all("span", class_="ticket-item__mmdd")  
                        dateEplus_beginning = None
                        dateEplus_ending = None

                        if date_year and date_mmdd:
                            for idx, (year, mmdd) in enumerate(zip(date_year, date_mmdd)):
                                year_text = year.get_text(strip=True)
                                mmdd_text = mmdd.get_text(strip=True)
                                if idx == 0:
                                    dateEplus_beginning = f"{year_text}{mmdd_text}"
                                if idx == len(date_mmdd) - 1:
                                    dateEplus_ending = f"{year_text}{mmdd_text}"
                                
                        div_eplus_venue = ticket.find("div", class_="ticket-item__venue")
                        if div_eplus_venue:
                            place_eplus = div_eplus_venue.find("p")
                            placeEplus = place_eplus.get_text(strip=True) 
                            linkEplus = "https://eplus.jp" + ticket.get("href") if ticket else None
                            if nameEplus and linkEplus:
                                Eplusconcerts.append({"Name": nameEplus, "Romaji": romajiEplus, "Place": placeEplus, "Date_beginning": dateEplus_beginning, "Date_ending": dateEplus_ending, "Link": linkEplus})
                                i+=1
                                retries = 0
                                #if i>1:
                                #    break #tester
                                #print(i)
                except Exception as e:
                    print(f"Error scraping from month: {month} page {page}: {e}")
                    retries -= 1
                    if retries > 0:
                        wait_time = random.randint(20, 60)
                        print(f"Retrying page {page} after {wait_time} seconds...")
                        time.sleep(wait_time)
                    else:
                        print(f"Failed to scrape page {page} after multiple attempts. Skipping.")
    print(f"Finished scraping current site. Proceeding to the next one.")
    return Eplusconcerts

def ltikeScrapper(doc_ltike):
    i=0
    ltikeconcerts = []
    pagination_position = doc_ltike.find("p", class_="Pagination__position")
    text = pagination_position.get_text(strip=True)
    match = re.search(r"（(\d+)ページ中）", text)
    if match:
        max_pages = int(match.group(1))
        print(f"Max pages: {max_pages}")
    
    for page in range(0, max_pages):
        url = f"https://l-tike.com/search/?keyword=*&area=3%2C5&pref=08%2C09%2C10%2C11%2C12%2C13%2C14%2C15%2C19%2C20%2C16%2C17%2C18%2C25%2C26%2C27%2C28%2C29%2C30&pdate_from=20250418&pdate_to=20250514&page={page}&ptabflg=0"
        print(f"Scraping page: {page+1}")
        
        retries = 5
        while retries > 0:
            try:
                doc_ltike = doc_from_url(url)
                tickets = doc_ltike.find_all("div", class_=["ResultBox boxContents prfSummaryItem", "ResultBox boxContents prfSummaryItem evenNumber"])
                for ticket in tickets:
                    nameltike = (ticket.find("h3", class_="ResultBox__title")).get_text(strip=True)
                
                    if nameltike:
                        romajiltike = convert_to_romaji(nameltike)
                    else:
                        romajiltike = None
                    info_block = ticket.find("dl", class_="ResultBox__informations")
                    dateltike = None
                    placeltike = None
                    if info_block:
                            # Find date and place from the <dl> block
                            date_block = info_block.find("div", class_="ResultBox__information")
                            if date_block and "公演日" in date_block.find("dt", class_="ResultBox__informationTitle").get_text(strip=True):
                                dateltike = date_block.find("dt", class_="ResultBox__informationText").get_text(strip=True)

                            place_block = info_block.find_all("div", class_="ResultBox__information")
                            for place in place_block:
                                if "会場" in place.find("dt", class_="ResultBox__informationTitle").get_text(strip=True):
                                    placeltike = place.find("dt", class_="ResultBox__informationText").get_text(strip=True)

                    
                    linkltike = "https://l-tike.com/search/?keyword=" + nameltike
                
                    if nameltike:
                        ltikeconcerts.append({"Name": nameltike, "Romaji": romajiltike, "Place": placeltike, "Date": dateltike, "Link": linkltike})
                        i+=1
                        retries = 0
                        #if i>1:
                        #    break #tester
                        #print(i)
                        #print (romajiltike, placeltike)
            except Exception as e:
                print(f"Error scraping page {page}: {e}")
                retries -= 1
                if retries > 0:
                    wait_time = random.randint(20, 60)
                    print(f"Retrying page {page} after {wait_time} seconds...")
                    time.sleep(wait_time)
                else:
                    print(f"Failed to scrape page {page} after multiple attempts. Skipping.")
    print(f"Finished scraping current site. Proceeding to the next one.")
    return ltikeconcerts

def OpenSheet(sheet_name):
    if os.path.exists(EXCEL_FILE):
        workbook = openpyxl.load_workbook(EXCEL_FILE)
    else:
        os.makedirs(BASE_FOLDER)
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)
    if sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
    else:
        sheet = workbook.create_sheet(title=sheet_name)
        sheet.append(HEADER) 
    sheet = workbook[sheet_name]
    return workbook, sheet

def remove_duplicates_in_excel_link(sheet_name="Events_t.pia.jp"):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    
    seen = {}
    row_number = 2
    while row_number <= sheet.max_row:
        link = sheet.cell(row=row_number, column=6).value
        if link in seen:
            sheet.delete_rows(row_number)
        else:
            seen[link] = row_number
            row_number += 1

    save_workbook(workbook)
    
def remove_duplicates_in_excel_name_place(sheet_name):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    
    seen = {}
    
    row_number = 2
    while row_number <= sheet.max_row:
        name = sheet.cell(row=row_number, column=1).value
        place = sheet.cell(row=row_number, column=3).value
        ending_date = sheet.cell(row=row_number, column=5).value
        
        identifier = (name, place)
        
        if identifier in seen:
            original_row = seen[identifier]
            original_ending_date = sheet.cell(row=original_row, column=5).value
            if ending_date > original_ending_date:
                sheet.cell(row=original_row, column=5).value = ending_date
            sheet.delete_rows(row_number)
        else:
            seen[identifier] = row_number
            row_number += 1
    
    save_workbook(workbook)
    
def splitter_pia(sheet_name):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    
    for row in range(2, sheet.max_row + 1):
        date_value = sheet.cell(row=row, column=4).value 
        
        if date_value and "～" in date_value:
            beginning_date, ending_date = date_value.split("～", 1)
            sheet.cell(row=row, column=4).value = beginning_date.strip()
            sheet.cell(row=row, column=5).value = ending_date.strip() 
        elif date_value:
            sheet.cell(row=row, column=4).value = date_value.strip()

    save_workbook(workbook)
    print(f"Finished splitting dates in {sheet_name}.")
    
def splitter_ltike(sheet_name):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    
    for row in range(2, sheet.max_row + 1):
        date_value = sheet.cell(row=row, column=4).value 
        
        if date_value:
            if "～" in date_value:
                beginning_date, ending_date = date_value.split("～", 1)
                sheet.cell(row=row, column=4).value = beginning_date.strip()
                sheet.cell(row=row, column=5).value = ending_date.strip()
            elif "・" in date_value:
                beginning_date, ending_date = date_value.split("・", 1)
                sheet.cell(row=row, column=4).value = beginning_date.strip()
                sheet.cell(row=row, column=5).value = ending_date.strip() 
            elif date_value:
                sheet.cell(row=row, column=4).value = date_value.strip() 

    save_workbook(workbook)
    print(f"Finished splitting dates in {sheet_name}.")
    
def cleaner(sheet_name):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]

    for row in range(2, sheet.max_row + 1):
        for column_index in [4, 5]:
            cell_value = sheet.cell(row=row, column=column_index).value
            if cell_value:
                try: 
                    cleaned_value = cell_value.split('(', 1)[0].strip()
                    date_value = datetime.strptime(cleaned_value, "%Y/%m/%d")
                    formatted_date = date_value.strftime("%Y-%m-%d")
                    sheet.cell(row=row, column=column_index).value = formatted_date
                except ValueError as e:
                    print(f"Row {row}, Column {column_index}: Invalid date '{cell_value}' - {e}")

    save_workbook(workbook)
    print(f"Finished cleaning dates in {sheet_name}.")

def style_sort_excel(sheet_name):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    sheet = workbook[sheet_name]
    row_count = sheet.max_row
    column_count = sheet.max_column
    
    df = pd.DataFrame(sheet.values)
    headers = df.iloc[0]
    df = df[1:]
    df.columns = headers
    df = df.sort_values(by="Beginning Date", ascending=True)
    sheet.delete_rows(1, sheet.max_row)
    for row in [df.columns.tolist()] + df.values.tolist():
       sheet.append(row)
    
    title_row_style = Font(size=14, color="FFFFFF", bold=True)
    for i in range (0,column_count):
        sheet.cell(row=1, column=i+1).font = title_row_style
    dim_holder = DimensionHolder(worksheet=sheet)
    for col in range(sheet.min_column, sheet.max_column + 1):
        if col == 6:
            dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=160)
        else:
            dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=35)
    sheet.column_dimensions = dim_holder
    
    for z in range (0, column_count):
        sheet.cell(row=1, column = z + 1).fill = PatternFill(start_color="38AA49", end_color="38AA49", fill_type="solid")
    for x in range(2, row_count):
        for z in range (0, column_count): 
            c = sheet.cell(row=x, column=z + 1)
            if x % 2 != 0:
                c.fill = PatternFill(start_color="ACFFB8", end_color="ACFFB8", fill_type="solid")
    if row_count % 2 != 0:
        for z in range (0, column_count):
            l = sheet.cell(row=row_count, column=z + 1)
            l.fill = PatternFill(start_color="ACFFB8", end_color="ACFFB8", fill_type="solid")
    
    save_workbook(workbook)
    
def combine_sheets(sheet_names):
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    combined_data = []
    
    for sheet_name in sheet_names:
        sheet = workbook[sheet_name]
        data = pd.DataFrame(sheet.values)
        data.columns = data.iloc[0]
        data = data.drop(0) 
        combined_data.append(data)
    combined_df = pd.concat(combined_data, ignore_index=True)
    if "Events_combined" in workbook.sheetnames:
        del workbook["Events_combined"]

    combined_sheet = workbook.create_sheet("Events_combined", 0)

    for r in dataframe_to_rows(combined_df, index=False, header=True):
        combined_sheet.append(r)
        
    save_workbook(workbook)
    style_sort_excel("Events_combined")
    print("Combined all sheets into 'Events_combined'.")
    
def OptiScrape_pia(url):
    doc_Pia = doc_from_url(url)
    return PiaScrapper(doc_Pia)

def pia_jp_scrap():
    ## Here we start scrapping pia.jp ##
    urls = [
        "https://t.pia.jp/music/",
        "https://t.pia.jp/anime/",
        "https://t.pia.jp/event/"
    ]
    Piaconcerts = []

    with concurrent.futures.ThreadPoolExecutor(max_workers=20) as executor:
        future_to_url = {executor.submit(OptiScrape_pia, url): url for url in urls}
        for future in concurrent.futures.as_completed(future_to_url):
            url = future_to_url[future]
            try:
                Piaconcerts.extend(future.result())
            except Exception as e:
                print(f"Error scraping {url}: {e}")

    sheet_name = "Events_t.pia.jp"
    workbook, sheet = OpenSheet(sheet_name)    
        
    for Piaconcert in Piaconcerts:
        sheet.append([Piaconcert["Name"], Piaconcert["Romaji"], Piaconcert["Place"], Piaconcert["Date"], Piaconcert["Date"], Piaconcert["Link"]])
        
    save_workbook(workbook)

    remove_duplicates_in_excel_link(sheet_name="Events_t.pia.jp")
    splitter_pia(sheet_name)
    cleaner(sheet_name)
    style_sort_excel(sheet_name)

    print(f"Done! Scraped t.pia.jp. Data saved to {EXCEL_FILE}.")
    
def eplus_jp_scrap():
    ## Here we start scrapping eplus.jp ##
    months = [4, 5]
    Eplusconcerts = []
    
    with ThreadPoolExecutor(max_workers=20) as executor:
        future_to_month = {executor.submit(eplusScrapper, month): month for month in months}
        
        for future in as_completed(future_to_month):
            month = future_to_month[future]
            try:
                Eplusconcert = future.result()
                Eplusconcerts.extend(Eplusconcert)
                print(f"Completed scraping for month {month}.")
            except Exception as e:
                print(f"Error scraping month {month}: {e}")

    sheet_name = "Events_eplus.jp"
    workbook, sheet = OpenSheet(sheet_name)

    for Eplusconcert in Eplusconcerts:
        sheet.append([Eplusconcert["Name"], Eplusconcert["Romaji"], Eplusconcert["Place"], Eplusconcert["Date_beginning"], Eplusconcert["Date_ending"], Eplusconcert["Link"]])
    save_workbook(workbook)

    remove_duplicates_in_excel_name_place(sheet_name)
    cleaner(sheet_name)
    style_sort_excel(sheet_name)

    print(f"Done! Scraped eplus.jp. Data saved to {EXCEL_FILE}.")
    
def ltike_jp_scrap():
    # Here we start scrapping l-tike ##

    doc_ltike_search = doc_from_url("https://l-tike.com/search/?keyword=*&area=3%2C5&pref=08%2C09%2C10%2C11%2C12%2C13%2C14%2C15%2C19%2C20%2C16%2C17%2C18%2C25%2C26%2C27%2C28%2C29%2C30&pdate_from=20250418&pdate_to=20250514&page=0&ptabflg=0")

    ltikeconcerts = ltikeScrapper(doc_ltike_search)

    sheet_name = "Events_l-tike.com"

    workbook, sheet = OpenSheet(sheet_name)

    for ltikeconcert in ltikeconcerts:
        sheet.append([ltikeconcert["Name"], ltikeconcert["Romaji"], ltikeconcert["Place"], ltikeconcert["Date"], ltikeconcert["Date"], ltikeconcert["Link"]])
        
    save_workbook(workbook)

    remove_duplicates_in_excel_name_place(sheet_name)
    splitter_ltike(sheet_name)
    cleaner(sheet_name)
    style_sort_excel(sheet_name)

    print(f"Done! Scraped l-tike.com. Data saved to {EXCEL_FILE}.")

sheet_names = []
pia = True
eplus = True
ltike = True

if not (pia or eplus or ltike):
    print("No websites selected. Exiting.")
else:    
    if pia:
        pia_jp_scrap()
        sheet_names.append("Events_t.pia.jp")
    if eplus:
        eplus_jp_scrap()
        sheet_names.append("Events_eplus.jp")
    if ltike:
        ltike_jp_scrap()
        sheet_names.append("Events_l-tike.com")

    if len(sheet_names) > 1:
        combine_sheets(sheet_names)